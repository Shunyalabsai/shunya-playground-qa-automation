import openpyxl
import json

data = {}

for fname in ['sheet1.xlsx', 'sheet2.xlsx']:
    wb = openpyxl.load_workbook(fname)
    data[fname] = {}
    for sname in wb.sheetnames:
        sheet = wb[sname]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
        sheet_rows = []
        for r in rows[1:]:
            if all(c is None for c in r):
                continue
            row_dict = {}
            for i, val in enumerate(r):
                if i < len(header):
                    row_dict[header[i]] = val
            sheet_rows.append(row_dict)
        data[fname][sname] = {
            'total_rows': len(sheet_rows),
            'header': header,
            'rows': sheet_rows
        }

with open('sheets_parsed.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2, default=str)

print("Parsed successfully!")
